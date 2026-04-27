import io
import zipfile
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


st.set_page_config(
    page_title="Consolidador de Empréstimos Consignados",
    page_icon="📊",
    layout="wide",
)

st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Sans:ital,wght@0,300;0,400;0,500;0,700;1,400&family=Syne:wght@600;700;800&display=swap');
    :root {
        --bg: #080808;
        --panel: #111111;
        --panel-2: #181818;
        --border: rgba(255, 255, 255, 0.07);
        --border-hover: rgba(255, 255, 255, 0.18);
        --text: #f0f0f0;
        --muted: #888888;
        --cyan: #00e5ff;
        --purple: #c77dff;
        --lime: #c8ff5d;
        --primary: #00e5ff;
        --primary-hover: #29ebff;
        --shadow: 0 20px 50px rgba(0,0,0,0.55);
        --hero-title-size: clamp(1.7rem, 3.8vw, 3rem);
    }
    html, body, [class*="css"]  {
        font-family: "DM Sans", sans-serif;
        scroll-behavior: smooth;
    }
    .main {
        position: relative;
        overflow: hidden;
        font-family: "DM Sans", sans-serif;
        background:
            radial-gradient(circle at 16% 6%, rgba(0,229,255,0.1), transparent 26%),
            radial-gradient(circle at 84% 85%, rgba(199,125,255,0.1), transparent 28%),
            linear-gradient(180deg, #0a0a0a 0%, #080808 100%);
        color: var(--text);
    }
    .main::before {
        content: "";
        position: fixed;
        inset: 0;
        pointer-events: none;
        opacity: 0.04;
        z-index: 0;
        background-image:
            radial-gradient(circle at 20% 20%, rgba(255,255,255,0.8) 1px, transparent 1px),
            radial-gradient(circle at 80% 75%, rgba(255,255,255,0.65) 1px, transparent 1px);
        background-size: 3px 3px, 4px 4px;
    }
    .bg-blob {
        position: fixed;
        filter: blur(120px);
        opacity: 0.12;
        z-index: 0;
        pointer-events: none;
        border-radius: 999px;
    }
    .blob-cyan { width: 600px; height: 600px; top: -230px; left: -180px; background: var(--cyan); }
    .blob-pink { width: 500px; height: 500px; right: -160px; bottom: -220px; background: var(--purple); }
    .blob-lime { width: 400px; height: 400px; left: 50%; top: 38%; transform: translate(-50%, -50%); background: var(--lime); }
    div[data-testid="stVerticalBlock"]:has(.hero-shell-anchor) {
        border: 1px solid var(--border);
        border-radius: 24px;
        padding: 1.2rem 1.35rem 1.35rem;
        margin-bottom: 1rem;
        background: linear-gradient(135deg, rgba(17,17,17,0.95), rgba(24,24,24,0.94));
        color: var(--text);
        backdrop-filter: blur(14px);
        box-shadow: var(--shadow);
        position: relative;
        z-index: 1;
    }
    .hero-head h1 {
        font-family: "Syne", sans-serif;
        font-size: var(--hero-title-size) !important;
        letter-spacing: -0.03em;
        line-height: 0.98;
        margin: 0 0 0.35rem 0;
        font-weight: 800;
    }
    .hero-head p {
        margin: 0 0 0.75rem;
        color: var(--muted);
        max-width: 56ch;
        font-size: 1.05rem;
    }
    .hero-divider {
        margin: 1rem 0 0.5rem 0;
        border: 0;
        border-top: 1px solid var(--border);
    }
    .hero-title {
        display: flex;
        align-items: flex-start;
        flex-direction: column;
        gap: 0.3rem;
    }
    .hero-icon {
        width: 26px;
        height: 26px;
        fill: none;
        stroke: var(--cyan);
        stroke-width: 2;
        stroke-linecap: round;
        stroke-linejoin: round;
        flex-shrink: 0;
        filter: drop-shadow(0 0 14px rgba(0,229,255,0.45));
    }
    .hero-highlight {
        background: linear-gradient(90deg, var(--cyan) 0%, #ff6ad5 100%);
        -webkit-background-clip: text;
        background-clip: text;
        color: transparent;
    }
    .hero-badge {
        display:inline-flex;
        align-items:center;
        gap:8px;
        padding:7px 12px;
        border-radius:999px;
        border:1px solid rgba(200,255,93,0.5);
        background:rgba(200,255,93,0.08);
        color:#ddff92;
        font-size:12px;
        font-family:"Syne",sans-serif;
        font-weight:600;
        letter-spacing:0.12em;
        text-transform:uppercase;
        margin-bottom:14px;
    }
    .fade-up { opacity: 0; transform: translateY(16px); animation: fadeUp .75s ease forwards; }
    .delay-1 { animation-delay: .08s; }
    .delay-2 { animation-delay: .18s; }
    .delay-3 { animation-delay: .28s; }
    .delay-4 { animation-delay: .38s; }
    @keyframes fadeUp { to { opacity: 1; transform: translateY(0); } }
    .feature-grid {
        display:grid;
        grid-template-columns:repeat(3,minmax(160px,1fr));
        gap:12px;
        min-width:300px;
    }
    .feature-card {
        border:1px solid var(--border);
        border-radius:20px;
        background:#121212;
        transition:transform .3s ease,border-color .3s ease;
        overflow:hidden;
    }
    .feature-card:hover { transform:translateY(-4px); border-color:var(--border-hover); }
    .feature-preview {
        height:88px;
        display:flex;
        align-items:center;
        justify-content:center;
        font-size:2rem;
        position:relative;
        background: radial-gradient(circle at 50% 45%, rgba(255,255,255,0.12) 0%, transparent 58%);
    }
    .feature-preview::before{
        content:"";
        position:absolute;
        width:74px;
        height:74px;
        border-radius:999px;
        filter:blur(22px);
        opacity:.38;
    }
    .preview-cyan::before{ background:var(--cyan);}
    .preview-purple::before{ background:var(--purple);}
    .preview-lime::before{ background:var(--lime);}
    .feature-body{ padding:10px 12px 12px; }
    .feature-body span { display:block;color:var(--muted);font-size:11px;letter-spacing:.08em;text-transform:uppercase; }
    .feature-body strong { font-family:"Syne",sans-serif;font-size:1rem;color:var(--text); }
    div[data-testid="stMetric"] {
        border: 1px solid var(--border);
        background-color: rgba(255,255,255,0.03);
        border-radius: 12px; padding: 0.55rem 0.7rem;
        box-shadow: 0 8px 20px rgba(0,0,0,0.25);
    }
    div[data-testid="stMetricLabel"] p { font-weight: 600; color: var(--muted); }
    div[data-testid="stMetricValue"] { color: var(--text); }
    .stDataFrame { border: 1px solid var(--border); border-radius: 14px; overflow: hidden; }
    .stDownloadButton button {
        width: 100%;
        border-radius: 14px; font-weight: 700; background: var(--primary); color: #001318;
        border: 1px solid var(--primary);
    }
    .stDownloadButton button:hover { background: var(--primary-hover); border-color: var(--primary-hover); }
    .stButton button { border-radius: 14px; font-weight: 700; border: 1px solid var(--border); transition: all .3s ease; }
    .stButton button:hover { border-color: var(--border-hover); transform: translateY(-2px); }
    @media (max-width: 600px) {
        .feature-grid { grid-template-columns:1fr; }
        .hero-head p { font-size: .95rem; }
    }
    .stSuccess, .stWarning, .stInfo { border-radius: 12px; }
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    """
    <div class="bg-blob blob-cyan"></div>
    <div class="bg-blob blob-pink"></div>
    <div class="bg-blob blob-lime"></div>
    <div class="hero-shell-anchor"></div>
    <div class="hero-head">
        <div style="display:flex;justify-content:space-between;gap:16px;align-items:flex-start;flex-wrap:wrap;">
            <div style="max-width:760px;">
                <div class="hero-badge fade-up">Consolidação inteligente</div>
                <h1 class="hero-title fade-up delay-1">
                    <svg class="hero-icon" viewBox="0 0 24 24" aria-hidden="true">
                        <path d="M3 3v18h18"></path>
                        <path d="M7 14l3-3 3 2 4-5"></path>
                        <circle cx="7" cy="14" r="1"></circle>
                        <circle cx="10" cy="11" r="1"></circle>
                        <circle cx="13" cy="13" r="1"></circle>
                        <circle cx="17" cy="8" r="1"></circle>
                    </svg>
                    <span>Consolidador de Empréstimos Consignados</span>
                    <span class="hero-highlight">CSV para Excel</span>
                </h1>
                <p class="fade-up delay-2">Envie ZIP ou múltiplos CSVs, consolide em segundos e exporte no layout do relatório.</p>
            </div>
            <div class="feature-grid">
                <div class="feature-card fade-up delay-2">
                    <div class="feature-preview preview-cyan">📄</div>
                    <div class="feature-body"><span>Formato ↗</span><strong>.CSV/.ZIP</strong></div>
                </div>
                <div class="feature-card fade-up delay-3">
                    <div class="feature-preview preview-purple">⚡</div>
                    <div class="feature-body"><span>Saída ↗</span><strong>.XLSX</strong></div>
                </div>
                <div class="feature-card fade-up delay-4">
                    <div class="feature-preview preview-lime">🟢</div>
                    <div class="feature-body"><span>Status ↗</span><strong>Pronto</strong></div>
                </div>
            </div>
        </div>
    </div>
    <hr class="hero-divider"/>
    """,
    unsafe_allow_html=True,
)

with st.sidebar:
    st.header("ℹ️ Informações")
    st.markdown(
        """
**Como usar:**
1. Faça upload de um arquivo ZIP com CSVs ou selecione múltiplos CSVs
2. Defina separador e codificação
3. (Opcional) informe a ordem final das colunas
4. Clique em **Consolidar Arquivos**
5. Baixe o Excel consolidado

**Formatos aceitos:**
- Arquivos CSV (`.csv`)
- Arquivos ZIP contendo CSVs (`.zip`)
"""
    )

st.header("1️⃣ Upload dos Arquivos")
upload_option = st.radio(
    "Escolha o tipo de upload:",
    ["Arquivo ZIP", "Múltiplos arquivos CSV"],
    horizontal=True,
)

st.header("2️⃣ Configurações")
col1, col2 = st.columns(2)

with col1:
    separador = st.selectbox(
        "Separador do CSV:",
        [";", ",", "\t", "|"],
        index=0,
        help="Caractere que separa as colunas nos CSVs",
    )

with col2:
    encoding = st.selectbox(
        "Codificação do arquivo:",
        ["utf-8", "latin1", "iso-8859-1", "cp1252"],
        index=0,
        help="Codificação de texto usada nos arquivos",
    )

create_bank_column = st.checkbox(
    "Criar coluna de banco no formato `codigo - descricao`",
    value=True,
    help=(
        "Quando existirem as colunas `ifConcessora.codigo` e "
        "`ifConcessora.descricao`, será criada a coluna `ifConcessora`."
    ),
)

st.header("3️⃣ Ordem das Colunas (Opcional)")
column_order_text = ""
preferred_columns = []

default_report_columns = [
    "matricula",
    "nomeTrabalhador",
    "inscricaoEmpregador.codigo",
    "nomeEmpregador",
    "valorParcela",
    "valorEmprestimo",
    "totalParcelas",
    "dataInicioContrato",
    "dataFimContrato",
    "competenciaInicioDesconto",
    "ifConcessora",
    "contrato",
    "inscricaoEmpregador.descricao",
    "numeroInscricaoEmpregador",
    "cpf",
    "competenciaFimDesconto",
    "valorLiberado",
    "qtdPagamentos",
    "qtdEscrituracoes",
    "categoriaTrabalhador.codigo",
    "categoriaTrabalhador.descricao",
    "competencia",
    "inscricaoEstabelecimento.codigo",
    "inscricaoEstabelecimento.descricao",
    "numeroInscricaoEstabelecimento",
    "dataAdmissao",
    "arquivo_origem",
]

use_default_report_order = st.checkbox(
    "Usar ordem padrão de colunas do relatório",
    value=True,
    help="Aplica automaticamente a ordem de colunas solicitada para o arquivo final.",
)

if not use_default_report_order:
    column_order_text = st.text_area(
        "Informe a ordem desejada das colunas (uma por linha)",
        placeholder="Exemplo:\nCPF\nNOME\nVALOR\nDATA",
        help=(
            "As colunas informadas serão posicionadas primeiro e na ordem exata definida. "
            "As demais colunas serão adicionadas automaticamente ao final."
        ),
    )
    preferred_columns = [line.strip() for line in column_order_text.splitlines() if line.strip()]

hide_columns_after_contract = st.checkbox(
    "Ocultar no Excel as colunas após `contrato`",
    value=True,
    help=(
        "Mantém visíveis apenas as 12 primeiras colunas do relatório "
        "(de `matricula` até `contrato`)."
    ),
)

uploaded_files = None


def _aplicar_estilo_planilha(worksheet):
    header_fill = PatternFill(fill_type="solid", start_color="1F2937", end_color="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    worksheet.freeze_panes = "A2"

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    for column_idx in range(1, worksheet.max_column + 1):
        column_letter = get_column_letter(column_idx)
        max_length = 0
        for row_idx in range(1, worksheet.max_row + 1):
            value = worksheet.cell(row=row_idx, column=column_idx).value
            value_length = len(str(value)) if value is not None else 0
            if value_length > max_length:
                max_length = value_length
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 45)


def _aplicar_filtros_planilha(worksheet, dataframe):
    worksheet.auto_filter.ref = worksheet.dimensions

    coluna_competencia = "competencia"
    coluna_competencia_inicio = "competenciaInicioDesconto"

    if (
        coluna_competencia in dataframe.columns
        and coluna_competencia_inicio in dataframe.columns
    ):
        valores_competencia = (
            dataframe[coluna_competencia].dropna().astype(str).str.strip()
        )
        valores_competencia = valores_competencia[valores_competencia != ""]

        if not valores_competencia.empty:
            competencia_alvo = valores_competencia.iloc[0]
            col_id = list(dataframe.columns).index(coluna_competencia_inicio)
            worksheet.auto_filter.add_filter_column(col_id, [competencia_alvo])

if upload_option == "Arquivo ZIP":
    uploaded_file = st.file_uploader(
        "Faça upload do arquivo ZIP contendo os CSVs:",
        type=["zip"],
        help="Selecione um ZIP que contenha os CSVs",
    )

    if uploaded_file:
        try:
            with zipfile.ZipFile(uploaded_file, "r") as zip_ref:
                csv_files = sorted(
                    [f for f in zip_ref.namelist() if f.lower().endswith(".csv")]
                )

                if not csv_files:
                    st.warning("⚠️ Nenhum arquivo CSV foi encontrado dentro do ZIP.")
                else:
                    st.success(f"✅ {len(csv_files)} arquivos CSV encontrados no ZIP")
                    uploaded_files = [
                        {"name": csv_file, "content": zip_ref.read(csv_file)}
                        for csv_file in csv_files
                    ]

        except Exception as error:
            st.error(f"❌ Erro ao processar o arquivo ZIP: {error}")
else:
    uploaded_csvs = st.file_uploader(
        "Faça upload dos arquivos CSV:",
        type=["csv"],
        accept_multiple_files=True,
        help="Selecione um ou mais CSVs para consolidar",
    )

    if uploaded_csvs:
        st.success(f"✅ {len(uploaded_csvs)} arquivo(s) CSV selecionado(s)")
        uploaded_files = [
            {"name": csv_file.name, "content": csv_file.getvalue()} for csv_file in uploaded_csvs
        ]

if uploaded_files:
    st.header("4️⃣ Consolidação")

    with st.expander("📋 Ver lista de arquivos"):
        for index, file_info in enumerate(uploaded_files, start=1):
            st.text(f"{index}. {file_info['name']}")

    if st.button("🚀 Consolidar Arquivos", type="primary", use_container_width=True):
        try:
            with st.spinner("Processando arquivos..."):
                dataframes = []
                progress_bar = st.progress(0)
                status_text = st.empty()

                for idx, file_info in enumerate(uploaded_files):
                    status_text.text(f"Processando: {file_info['name']}")
                    df = pd.read_csv(
                        io.BytesIO(file_info["content"]),
                        sep=separador,
                        encoding=encoding,
                    )
                    df["arquivo_origem"] = file_info["name"]
                    dataframes.append(df)
                    progress_bar.progress((idx + 1) / len(uploaded_files))

                status_text.text("Consolidando dados...")
                df_consolidado = pd.concat(dataframes, ignore_index=True, sort=False)

                if "inscricaoEmpregador.codigo" in df_consolidado.columns:
                    df_consolidado["inscricaoEmpregador.codigo"] = ""

                if create_bank_column:
                    bank_code_column = "ifConcessora.codigo"
                    bank_description_column = "ifConcessora.descricao"

                    if (
                        bank_code_column in df_consolidado.columns
                        and bank_description_column in df_consolidado.columns
                    ):
                        bank_code = (
                            df_consolidado[bank_code_column]
                            .astype("string")
                            .fillna("")
                            .str.strip()
                            .str.replace(r"\\.0+$", "", regex=True)
                        )
                        bank_description = (
                            df_consolidado[bank_description_column]
                            .astype("string")
                            .fillna("")
                            .str.strip()
                        )

                        one_digit_mask = bank_code.str.fullmatch(r"\\d")
                        bank_code = bank_code.where(~one_digit_mask, bank_code.str.zfill(2))

                        combined_bank = (bank_code + " - " + bank_description).str.strip()
                        combined_bank = combined_bank.str.replace(r"^\\s*-\\s*", "", regex=True)
                        combined_bank = combined_bank.str.replace(r"\\s*-\\s*$", "", regex=True)
                        df_consolidado["ifConcessora"] = combined_bank
                    else:
                        st.warning(
                            "⚠️ Não foi possível criar `ifConcessora`: "
                            "colunas `ifConcessora.codigo` e/ou "
                            "`ifConcessora.descricao` não encontradas."
                        )

                if use_default_report_order:
                    existing_default_columns = [
                        column for column in default_report_columns if column in df_consolidado.columns
                    ]
                    missing_default_columns = [
                        column for column in default_report_columns if column not in df_consolidado.columns
                    ]
                    remaining_columns = [
                        column
                        for column in df_consolidado.columns
                        if column not in existing_default_columns
                    ]
                    final_columns = existing_default_columns + remaining_columns
                    df_consolidado = df_consolidado[final_columns]

                    if missing_default_columns:
                        st.warning(
                            "⚠️ Algumas colunas da ordem padrão não foram encontradas: "
                            + ", ".join(missing_default_columns)
                        )
                elif preferred_columns:
                    existing_preferred = [
                        column for column in preferred_columns if column in df_consolidado.columns
                    ]
                    missing_preferred = [
                        column for column in preferred_columns if column not in df_consolidado.columns
                    ]
                    remaining_columns = [
                        column
                        for column in df_consolidado.columns
                        if column not in existing_preferred
                    ]
                    final_columns = existing_preferred + remaining_columns
                    df_consolidado = df_consolidado[final_columns]

                    if missing_preferred:
                        st.warning(
                            "⚠️ Algumas colunas informadas não foram encontradas: "
                            + ", ".join(missing_preferred)
                        )

                progress_bar.empty()
                status_text.empty()

                st.success("✅ Consolidação concluída com sucesso!")
                m1, m2, m3 = st.columns(3)
                m1.metric("Arquivos Processados", len(uploaded_files))
                m2.metric("Total de Linhas", len(df_consolidado))
                m3.metric("Total de Colunas", len(df_consolidado.columns))

                st.subheader("👀 Prévia dos Dados Consolidados")
                st.dataframe(df_consolidado.head(10), use_container_width=True)

                with st.expander("📊 Informações das Colunas"):
                    st.write(f"**Colunas encontradas ({len(df_consolidado.columns)}):**")
                    for column in df_consolidado.columns:
                        st.text(f"• {column}")

                output = io.BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    df_consolidado.to_excel(writer, index=False, sheet_name="Dados Consolidados")
                    worksheet = writer.sheets["Dados Consolidados"]
                    _aplicar_filtros_planilha(worksheet, df_consolidado)
                    _aplicar_estilo_planilha(worksheet)
                    if hide_columns_after_contract:
                        first_hidden_column = 13
                        for column_idx in range(first_hidden_column, worksheet.max_column + 1):
                            column_letter = get_column_letter(column_idx)
                            worksheet.column_dimensions[column_letter].hidden = True

                output_filename = f"consolidado_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                st.download_button(
                    label="📥 Baixar Planilha Consolidada (Excel)",
                    data=output.getvalue(),
                    file_name=output_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

        except Exception as error:
            st.error(f"❌ Erro ao processar os arquivos: {error}")
            st.exception(error)
else:
    st.info("👆 Faça upload dos arquivos para começar")

st.markdown("---")
st.markdown(
    """
<div style='text-align: center'>
    <p>Desenvolvido com ❤️ usando Streamlit</p>
</div>
""",
    unsafe_allow_html=True,
)
